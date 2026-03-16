#!/usr/bin/env python3
"""
Excel Report Generator — Professional Financial Model
=====================================================
Generates a 12-sheet, formula-driven Excel workbook for migration roadmap analysis.

Design Principles:
  • All tunable parameters live in the Assumptions sheet
  • Derived values use Excel formulas — change an assumption, everything recalculates
  • Blue font = editable input, Black font = formula (financial-model convention)
  • Named ranges improve readability of cross-sheet formulas
  • Excel Tables with structured references for data sheets
  • Conditional formatting for heat-maps, data-bars, icon-sets
  • Charts linked to formula cells
"""

import json
from typing import Dict, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, GradientFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, CellIsRule
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


# ════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ════════════════════════════════════════════════════════════════════════════

COLORS = {
    'primary':          '17324D',
    'primary_light':    '4F7DA4',
    'secondary':        '22303C',
    'accent_green':     '2E8B57',
    'accent_amber':     'D4A72C',
    'accent_red':       'C8553D',
    'background':       'F5F7FA',
    'panel':            'FFFFFF',
    'panel_alt':        'ECF2F8',
    'white':            'FFFFFF',
    'border':           'D6DEE8',
    'text_muted':       '66788A',
    'input_blue':       '2F6FED',
    'formula_black':    '22303C',
    'cross_ref_green':  '1F7A4D',
    'assumption_bg':    'FFF4CC',
    # Mode colours (light fills)
    'build_to_legacy':  'FBE4DF',
    'bridge_to_model':  'E0E7F6',
    'strategic':        'DFF0E5',
    # Mode colours (solid for charts / Gantt)
    'mode_btl':         'D97B67',
    'mode_btm':         '6A7DB3',
    'mode_str':         '4D9A6B',
    'quarter_fill':     'DCEAF7',
    'card_accent':      '86A7C2',
}

THIN_BORDER = Border(
    left=Side(style='thin', color=COLORS['border']),
    right=Side(style='thin', color=COLORS['border']),
    top=Side(style='thin', color=COLORS['border']),
    bottom=Side(style='thin', color=COLORS['border']),
)

CARD_BORDER = Border(
    left=Side(style='thin', color=COLORS['border']),
    right=Side(style='thin', color=COLORS['border']),
    top=Side(style='medium', color=COLORS['card_accent']),
    bottom=Side(style='thin', color=COLORS['border']),
)

FONT_INPUT = Font(name='Aptos', size=11, color=COLORS['input_blue'], bold=True)
FONT_FORMULA = Font(name='Aptos', size=11, color=COLORS['formula_black'])
FONT_HEADER = Font(name='Aptos', size=10, bold=True, color=COLORS['white'])
FONT_SECTION = Font(name='Aptos', size=12, bold=True, color=COLORS['primary'])
FONT_TITLE = Font(name='Aptos Display', size=20, bold=True, color=COLORS['white'])
FONT_SUBTITLE = Font(name='Aptos', size=10, italic=True, color=COLORS['text_muted'])
FONT_LABEL = Font(name='Aptos', size=10, color=COLORS['secondary'])
FONT_CARD_VALUE = Font(name='Aptos Display', size=19, bold=True, color=COLORS['primary'])
FONT_CARD_LABEL = Font(name='Aptos', size=9, bold=True, color=COLORS['text_muted'])
FONT_CARD_NOTE = Font(name='Aptos', size=9, italic=True, color=COLORS['text_muted'])

FILL_HEADER = PatternFill('solid', fgColor=COLORS['primary'])
FILL_SECTION = PatternFill('solid', fgColor=COLORS['panel_alt'])
FILL_ASSUMPTION = PatternFill('solid', fgColor=COLORS['assumption_bg'])
FILL_BACKGROUND = PatternFill('solid', fgColor=COLORS['background'])
FILL_WHITE = PatternFill('solid', fgColor=COLORS['white'])
FILL_PANEL = PatternFill('solid', fgColor=COLORS['panel'])
FILL_PANEL_ALT = PatternFill('solid', fgColor=COLORS['panel_alt'])

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')

SCENARIO_KEYS  = ['FastExit', 'Balanced', 'TargetFirst']
SCENARIO_NAMES = ['Fast Exit', 'Balanced', 'Target-First']
SCHEDULE_SHEETS = [f'Schedule_{k}' for k in SCENARIO_KEYS]
MODE_NAMES = ['Build-to-Legacy', 'Bridge-to-Model', 'Strategic']


# ════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════════════════════

def _hdr_cell(ws, row, col, value, width=None):
    """Write a header cell with standard formatting."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _input_cell(ws, row, col, value, fmt=None):
    """Write an editable input cell (blue font, yellow bg)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_INPUT
    c.fill = FILL_ASSUMPTION
    c.border = THIN_BORDER
    c.alignment = ALIGN_RIGHT
    if fmt:
        c.number_format = fmt
    return c


def _formula_cell(ws, row, col, formula, fmt=None):
    """Write a formula cell (black font)."""
    c = ws.cell(row=row, column=col, value=formula)
    c.font = FONT_FORMULA
    c.border = THIN_BORDER
    c.alignment = ALIGN_RIGHT
    if fmt:
        c.number_format = fmt
    return c


def _label_cell(ws, row, col, value, bold=False):
    """Write a label cell."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Aptos', size=10, bold=bold, color=COLORS['secondary'])
    c.border = THIN_BORDER
    c.alignment = ALIGN_LEFT
    c.fill = FILL_PANEL
    return c


def _section_header(ws, row, col_start, col_end, title):
    """Merge cells and write a soft section header bar."""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.font = FONT_SECTION
    c.fill = FILL_SECTION
    c.alignment = ALIGN_LEFT
    c.border = Border(
        left=Side(style='medium', color=COLORS['primary']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border']),
    )
    for ci in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=ci).fill = FILL_SECTION
        ws.cell(row=row, column=ci).border = Border(
            top=Side(style='thin', color=COLORS['border']),
            right=Side(style='thin', color=COLORS['border']),
            bottom=Side(style='thin', color=COLORS['border']),
        )


def _sheet_setup(ws):
    """Apply workbook-wide sheet defaults."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90


def _sheet_banner(ws, col_start, col_end, title, subtitle='', accent_color=None):
    """Create a consistent two-row banner for each sheet."""
    _sheet_setup(ws)
    accent = accent_color or COLORS['primary_light']
    gradient = GradientFill(
        type='linear',
        degree=0,
        stop=(COLORS['primary'], accent),
    )

    ws.merge_cells(start_row=1, start_column=col_start,
                   end_row=1, end_column=col_end)
    for ci in range(col_start, col_end + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = gradient
        cell.border = Border(
            bottom=Side(style='medium', color=accent),
        )
    c = ws.cell(row=1, column=col_start, value=title)
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=col_start,
                   end_row=2, end_column=col_end)
    for ci in range(col_start, col_end + 1):
        cell = ws.cell(row=2, column=ci)
        cell.fill = FILL_PANEL_ALT
        cell.border = Border(
            bottom=Side(style='thin', color=COLORS['border']),
        )
    c2 = ws.cell(row=2, column=col_start, value=subtitle)
    c2.font = FONT_SUBTITLE
    c2.alignment = ALIGN_LEFT
    ws.row_dimensions[2].height = 20


def _metric_card(ws, row, col_start, col_end, value, label, fmt=None,
                 accent_color=None, note=None):
    """Render a dashboard-style metric card."""
    accent = accent_color or COLORS['primary_light']

    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    ws.merge_cells(start_row=row + 1, start_column=col_start,
                   end_row=row + 1, end_column=col_end)
    if note:
        ws.merge_cells(start_row=row + 2, start_column=col_start,
                       end_row=row + 2, end_column=col_end)

    for r in range(row, row + (3 if note else 2)):
        for ci in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = FILL_PANEL
            cell.border = Border(
                left=Side(style='thin', color=COLORS['border']),
                right=Side(style='thin', color=COLORS['border']),
                top=Side(style='medium' if r == row else 'thin',
                         color=accent if r == row else COLORS['border']),
                bottom=Side(style='thin', color=COLORS['border']),
            )

    value_cell = ws.cell(row=row, column=col_start, value=value)
    value_cell.font = FONT_CARD_VALUE
    value_cell.alignment = ALIGN_CENTER
    if fmt:
        value_cell.number_format = fmt

    label_cell = ws.cell(row=row + 1, column=col_start, value=label)
    label_cell.font = FONT_CARD_LABEL
    label_cell.alignment = ALIGN_CENTER

    if note:
        note_cell = ws.cell(row=row + 2, column=col_start, value=note)
        note_cell.font = FONT_CARD_NOTE
        note_cell.alignment = ALIGN_CENTER


def _callout_row(ws, row, col_start, col_end, text, accent_color=None):
    """Render a merged callout row used for recommendations and notes."""
    accent = accent_color or COLORS['accent_green']
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    for ci in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = FILL_PANEL
        cell.border = Border(
            left=Side(style='medium', color=accent if ci == col_start else COLORS['border']),
            right=Side(style='thin', color=COLORS['border']),
            top=Side(style='thin', color=COLORS['border']),
            bottom=Side(style='thin', color=COLORS['border']),
        )
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = FONT_FORMULA
    c.alignment = ALIGN_LEFT


def _col_letter(col_index):
    """1-based column index → letter."""
    return get_column_letter(col_index)


def _define_name(wb, name, sheet_title, cell_ref):
    """Create a workbook-scoped named range."""
    ref = f"'{sheet_title}'!{cell_ref}"
    try:
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)
    except Exception:
        pass  # skip duplicate


# ════════════════════════════════════════════════════════════════════════════
# MAIN GENERATOR CLASS
# ════════════════════════════════════════════════════════════════════════════

class ExcelReportGenerator:
    """Generates a 12-sheet, formula-driven Excel workbook."""

    def __init__(self, optimization_results: List[Dict],
                 packages: List[Dict], config: Dict):
        self.results = optimization_results
        self.packages = packages
        self.config = config
        self.wb = Workbook()
        # Remove the default sheet created by openpyxl
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Computed helpers
        self.n_packages = len(packages)
        self.scenario_schedules: Dict[str, List[Dict]] = {}
        self._prepare_schedules()

        # Cell address maps (populated during creation)
        self._assumption_cells: Dict[str, str] = {}
        # Row where package data starts in Package_Data
        self._pkg_data_start_row = 0
        self._pkg_data_end_row = 0
        # Row where schedule data starts per sheet
        self._sched_start_rows: Dict[str, int] = {}
        self._sched_end_rows: Dict[str, int] = {}

    # ── Preparation ────────────────────────────────────────────────────

    def _prepare_schedules(self):
        """Organise results into the three canonical scenarios."""
        for i, key in enumerate(SCENARIO_KEYS):
            if i < len(self.results):
                result = self.results[i]
                schedule = result.get('schedule', [])
                schedule_sorted = sorted(schedule,
                                         key=lambda x: x.get('start_month', 0))
                self.scenario_schedules[key] = schedule_sorted
            else:
                self.scenario_schedules[key] = []

    # ── Public entry point ─────────────────────────────────────────────

    def generate_report(self, output_path: str) -> str:
        """Generate the complete 12-sheet workbook."""
        print("Generating professional financial-model Excel report …")

        self._create_assumptions()
        self._create_package_data()
        for i, key in enumerate(SCENARIO_KEYS):
            self._create_schedule_sheet(key, SCENARIO_NAMES[i])
        self._create_executive_summary()
        self._create_scenario_comparison()
        self._create_gantt_timeline()
        self._create_effort_cost_model()
        self._create_risk_assessment()
        self._create_dependency_analysis()
        self._create_data_appendix()

        desired_order = [
            'Executive_Summary',
            'Scenario_Comparison',
            'Timeline_Gantt',
            'Effort_Cost_Model',
            'Risk_Assessment',
            'Dependency_Analysis',
            'Schedule_FastExit',
            'Schedule_Balanced',
            'Schedule_TargetFirst',
            'Assumptions',
            'Package_Data',
            'Data_Appendix',
        ]
        self.wb._sheets = [
            self.wb[name] for name in desired_order if name in self.wb.sheetnames
        ]
        self.wb.active = 0

        self.wb.save(output_path)
        print(f"Report saved → {output_path}")
        return output_path

    # ══════════════════════════════════════════════════════════════════
    # SHEET 1: ASSUMPTIONS
    # ══════════════════════════════════════════════════════════════════

    def _create_assumptions(self):
        ws = self.wb.create_sheet('Assumptions')
        ws.sheet_properties.tabColor = COLORS['accent_amber']
        _sheet_banner(
            ws, 2, 8,
            'Assumptions & Parameters',
            'Editable inputs are highlighted in gold. Change them here and the full workbook recalculates.',
            COLORS['accent_amber'],
        )

        # Column widths
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 5
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18

        row = 4
        # ── General Parameters ─────────────────────────────────────
        _section_header(ws, row, 2, 3, 'General Parameters')
        row += 1

        general_params = [
            ('Migration Horizon (months)',  'horizon_months',
             self.config.get('migration_horizon_months', 60), '#,##0'),
            ('Team Capacity (parallel packages)',  'team_capacity',
             self.config.get('team_capacity', 6), '#,##0'),
            ('FTE Daily Cost Rate ($)',  'fte_daily_rate',
             self.config.get('fte_daily_rate', 850), '$#,##0'),
            ('Working Days per Month',  'working_days_month', 20, '#,##0'),
            ('Risk Buffer Factor',  'risk_buffer_factor', 0.25, '0.00'),
            ('Dependency Criticality Threshold', 'dep_crit_threshold', 3, '#,##0'),
            ('Complexity Criticality Threshold', 'complexity_crit_threshold', 3.5, '0.0'),
        ]

        for label, name, value, fmt in general_params:
            _label_cell(ws, row, 2, label, bold=True)
            _input_cell(ws, row, 3, value, fmt)
            _define_name(self.wb, name, 'Assumptions',
                         f'$C${row}')
            self._assumption_cells[name] = f'Assumptions!$C${row}'
            row += 1

        row += 1

        # ── Mode Parameters Table ──────────────────────────────────
        _section_header(ws, row, 2, 6, 'Mode Parameters')
        row += 1
        mode_table_start = row

        headers = ['Mode', 'Duration Multiplier', 'Tech Debt Penalty',
                   'Resource Efficiency', 'Cost Multiplier']
        for ci, h in enumerate(headers, 2):
            _hdr_cell(ws, row, ci, h)

        row += 1
        mode_defaults = [
            ('Build-to-Legacy', 0.7, 3.0, 1.2, 0.8),
            ('Bridge-to-Model', 1.0, 1.5, 1.0, 1.0),
            ('Strategic',       1.4, 0.5, 0.8, 1.3),
        ]
        for mode_name, dur_mult, td_pen, res_eff, cost_mult in mode_defaults:
            _label_cell(ws, row, 2, mode_name, bold=True)
            _input_cell(ws, row, 3, dur_mult, '0.00')
            _input_cell(ws, row, 4, td_pen, '0.00')
            _input_cell(ws, row, 5, res_eff, '0.00')
            _input_cell(ws, row, 6, cost_mult, '0.00')
            row += 1

        mode_table_end = row - 1

        # Create an Excel Table for the mode parameters
        tbl_ref = f"B{mode_table_start}:F{mode_table_end}"
        tbl = Table(displayName="ModeTable", ref=tbl_ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium13", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl)

        # Named ranges for mode param columns
        _define_name(self.wb, 'mode_dur_mult', 'Assumptions',
                     f'$C${mode_table_start+1}:$C${mode_table_end}')
        _define_name(self.wb, 'mode_names_rng', 'Assumptions',
                     f'$B${mode_table_start+1}:$B${mode_table_end}')
        self._assumption_cells['mode_table_start'] = mode_table_start
        self._assumption_cells['mode_table_end'] = mode_table_end

        row += 1

        # ── Domain Priority Weights ────────────────────────────────
        _section_header(ws, row, 2, 4, 'Domain Priority & Earliest Start')
        row += 1
        dom_table_start = row

        _hdr_cell(ws, row, 2, 'Domain')
        _hdr_cell(ws, row, 3, 'Priority Weight')
        _hdr_cell(ws, row, 4, 'Earliest Start (Month)')
        row += 1

        domains = sorted({p.get('domain', 'Unknown') for p in self.packages})
        if not domains:
            domains = ['Unknown']
        for domain in domains:
            _label_cell(ws, row, 2, domain, bold=True)
            _input_cell(ws, row, 3, 1.0, '0.00')
            _input_cell(ws, row, 4, 0, '#,##0')
            row += 1

        dom_table_end = row - 1
        tbl_ref2 = f"B{dom_table_start}:D{dom_table_end}"
        tbl2 = Table(displayName="DomainTable", ref=tbl_ref2)
        tbl2.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium13", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl2)

        row += 1

        # ── Scenario Objective Weights ─────────────────────────────
        _section_header(ws, row, 2, 6, 'Scenario Objective Weights')
        row += 1
        scen_table_start = row

        _hdr_cell(ws, row, 2, 'Scenario')
        _hdr_cell(ws, row, 3, 'Duration Weight')
        _hdr_cell(ws, row, 4, 'Strategic Weight')
        _hdr_cell(ws, row, 5, 'Tech Debt Weight')
        row += 1

        scen_weights = [
            ('Fast Exit',     0.7, 0.1, 0.2),
            ('Balanced',      0.33, 0.34, 0.33),
            ('Target-First',  0.1, 0.6, 0.3),
        ]
        for sname, dw, sw, tw in scen_weights:
            _label_cell(ws, row, 2, sname, bold=True)
            _input_cell(ws, row, 3, dw, '0.00')
            _input_cell(ws, row, 4, sw, '0.00')
            _input_cell(ws, row, 5, tw, '0.00')
            row += 1

        scen_table_end = row - 1
        tbl_ref3 = f"B{scen_table_start}:E{scen_table_end}"
        tbl3 = Table(displayName="ScenarioWeightsTable", ref=tbl_ref3)
        tbl3.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium13", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tbl3)

        row += 1

        # ── Risk Weights ───────────────────────────────────────────
        _section_header(ws, row, 2, 3, 'Risk Assessment Weights')
        row += 1

        risk_params = [
            ('Complexity Weight',   'risk_w_complexity',  0.35, '0.00'),
            ('Dependency Weight',   'risk_w_dependency',  0.25, '0.00'),
            ('Business Value Weight','risk_w_bv',         0.20, '0.00'),
            ('Volume Weight',       'risk_w_volume',      0.20, '0.00'),
        ]
        for label, name, value, fmt in risk_params:
            _label_cell(ws, row, 2, label, bold=True)
            _input_cell(ws, row, 3, value, fmt)
            _define_name(self.wb, name, 'Assumptions', f'$C${row}')
            self._assumption_cells[name] = f'Assumptions!$C${row}'
            row += 1

        ws.freeze_panes = 'C5'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 2: PACKAGE_DATA
    # ══════════════════════════════════════════════════════════════════

    def _create_package_data(self):
        ws = self.wb.create_sheet('Package_Data')
        ws.sheet_properties.tabColor = COLORS['primary']
        _sheet_banner(
            ws, 1, 10,
            'Package Data Hub',
            'Master package register used by the schedules, dashboards, risk model and appendix tabs.',
            COLORS['primary_light'],
        )

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 14
        ws.column_dimensions['J'].width = 16

        headers = ['Package_ID', 'Name', 'Domain', 'Job_Count',
                   'Effort_Days', 'Complexity', 'Business_Value',
                   'Risk_Score', 'Upstream_Count', 'Downstream_Count']

        ws.merge_cells('A3:J3')
        note = ws['A3']
        note.value = 'Use this tab as the canonical package inventory. It is organised for fast filtering and downstream formula lookups.'
        note.font = FONT_SUBTITLE
        note.alignment = ALIGN_LEFT

        hdr_row = 4
        for ci, h in enumerate(headers, 1):
            _hdr_cell(ws, hdr_row, ci, h)

        start_row = 5
        for ri, pkg in enumerate(self.packages, start_row):
            ws.cell(row=ri, column=1, value=pkg.get('package_id',
                    pkg.get('id', f'PKG-{ri-1:03d}'))).font = FONT_FORMULA
            ws.cell(row=ri, column=2,
                    value=pkg.get('name', '')).font = FONT_FORMULA
            ws.cell(row=ri, column=3,
                    value=pkg.get('domain', '')).font = FONT_FORMULA
            ws.cell(row=ri, column=4,
                    value=pkg.get('job_count',
                                  len(pkg.get('job_ids', [])))).font = FONT_FORMULA
            ws.cell(row=ri, column=5,
                    value=pkg.get('total_effort_days',
                                  pkg.get('effort_days', 0))).font = FONT_FORMULA
            ws.cell(row=ri, column=6,
                    value=pkg.get('complexity_score',
                                  pkg.get('complexity', 0))).font = FONT_FORMULA
            ws.cell(row=ri, column=7,
                    value=pkg.get('business_value', 0)).font = FONT_FORMULA
            ws.cell(row=ri, column=8,
                    value=pkg.get('risk_score', 0)).font = FONT_FORMULA
            ws.cell(row=ri, column=9,
                    value=pkg.get('upstream_count',
                                  len(pkg.get('upstream_packages', [])))).font = FONT_FORMULA
            ws.cell(row=ri, column=10,
                    value=pkg.get('downstream_count',
                                  len(pkg.get('downstream_packages', [])))).font = FONT_FORMULA

            for ci in range(1, 11):
                ws.cell(row=ri, column=ci).border = THIN_BORDER
                if ci >= 4:
                    ws.cell(row=ri, column=ci).number_format = '#,##0'
                    ws.cell(row=ri, column=ci).alignment = ALIGN_RIGHT
                if ci == 6:
                    ws.cell(row=ri, column=ci).number_format = '0.0'
                if ci == 8:
                    ws.cell(row=ri, column=ci).number_format = '0.0'

        end_row = start_row + self.n_packages - 1
        self._pkg_data_start_row = start_row
        self._pkg_data_end_row = end_row

        # Create Excel Table
        if self.n_packages > 0:
            tbl_ref = f"A{hdr_row}:J{end_row}"
            tbl = Table(displayName="PackageData", ref=tbl_ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium13", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True,
                showColumnStripes=False)
            ws.add_table(tbl)

        # Data bars on Complexity
        if self.n_packages > 0:
            ws.conditional_formatting.add(
                f'F{start_row}:F{end_row}',
                DataBarRule(start_type='min', end_type='max',
                            color='5C8AB8', showValue=True,
                            minLength=None, maxLength=None))

        # Color scale on Risk Score
        if self.n_packages > 0:
            ws.conditional_formatting.add(
                f'H{start_row}:H{end_row}',
                ColorScaleRule(start_type='min', start_color='28A745',
                               mid_type='percentile', mid_value=50,
                               mid_color='FFC107',
                               end_type='max', end_color='DC3545'))

        ws.auto_filter.ref = f"A{hdr_row}:J{end_row}"
        ws.freeze_panes = 'A5'

        # Named ranges for key columns
        _define_name(self.wb, 'pkg_ids', 'Package_Data',
                     f'$A${start_row}:$A${end_row}')
        _define_name(self.wb, 'pkg_effort', 'Package_Data',
                     f'$E${start_row}:$E${end_row}')
        _define_name(self.wb, 'pkg_complexity', 'Package_Data',
                     f'$F${start_row}:$F${end_row}')
        _define_name(self.wb, 'pkg_bv', 'Package_Data',
                     f'$G${start_row}:$G${end_row}')
        _define_name(self.wb, 'pkg_risk', 'Package_Data',
                     f'$H${start_row}:$H${end_row}')

    # ══════════════════════════════════════════════════════════════════
    # SHEETS 3-5: SCHEDULE (per scenario)
    # ══════════════════════════════════════════════════════════════════

    def _create_schedule_sheet(self, scenario_key: str, scenario_label: str):
        sheet_name = f'Schedule_{scenario_key}'
        ws = self.wb.create_sheet(sheet_name)

        # Tab colours
        tab_colors = {
            'FastExit': COLORS['accent_red'],
            'Balanced': COLORS['accent_amber'],
            'TargetFirst': COLORS['accent_green'],
        }
        accent_color = tab_colors.get(scenario_key, COLORS['primary_light'])
        ws.sheet_properties.tabColor = accent_color
        _sheet_banner(
            ws, 1, 9,
            f'{scenario_label} Scenario Schedule',
            'Package-by-package roadmap with timing, delivery mode and effort profile.',
            accent_color,
        )

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 17
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 14

        headers = ['Package_ID', 'Name', 'Domain', 'Start_Month',
                   'End_Month', 'Duration', 'Mode', 'Effort_Days',
                   'Business_Value']

        hdr_row = 8
        for ci, h in enumerate(headers, 1):
            _hdr_cell(ws, hdr_row, ci, h)

        schedule = self.scenario_schedules.get(scenario_key, [])

        start_row = hdr_row + 1

        for ri, item in enumerate(schedule, start_row):
            pkg_id = item.get('package_id', item.get('id', ''))
            pkg_name = item.get('package_name', item.get('name', ''))
            domain = item.get('domain', '')
            start_m = item.get('start_month', 0)
            end_m = item.get('end_month', 0)
            mode = item.get('selected_mode', item.get('mode', ''))
            # Normalise mode name for display
            mode_display = mode.replace('_', '-').title()
            if mode_display == 'Build-To-Legacy':
                mode_display = 'Build-to-Legacy'
            elif mode_display == 'Bridge-To-Model':
                mode_display = 'Bridge-to-Model'
            effort = item.get('effort_days', 0)
            bv = item.get('business_value', 0)

            ws.cell(row=ri, column=1, value=pkg_id).font = FONT_FORMULA
            ws.cell(row=ri, column=2, value=pkg_name).font = FONT_FORMULA
            ws.cell(row=ri, column=3, value=domain).font = FONT_FORMULA
            ws.cell(row=ri, column=4, value=start_m).font = FONT_FORMULA
            ws.cell(row=ri, column=4).number_format = '#,##0'
            ws.cell(row=ri, column=5, value=end_m).font = FONT_FORMULA
            ws.cell(row=ri, column=5).number_format = '#,##0'

            # Duration formula: =E{row}-D{row}  (simple difference)
            # Also add mode multiplier: =ROUND((H{row}/working_days)*VLOOKUP(G{row},ModeTable,2,FALSE),0)
            # We'll use the simpler version for the "display" duration and a
            # calculated duration as well.  Keep End-Start for clarity.
            dur_formula = f'=E{ri}-D{ri}'
            _formula_cell(ws, ri, 6, dur_formula, '#,##0')

            ws.cell(row=ri, column=7, value=mode_display).font = FONT_FORMULA
            ws.cell(row=ri, column=8, value=effort).font = FONT_FORMULA
            ws.cell(row=ri, column=8).number_format = '#,##0'
            ws.cell(row=ri, column=9, value=bv).font = FONT_FORMULA
            ws.cell(row=ri, column=9).number_format = '#,##0'

            for ci in range(1, 10):
                ws.cell(row=ri, column=ci).border = THIN_BORDER

        end_row = start_row + len(schedule) - 1 if schedule else start_row
        self._sched_start_rows[scenario_key] = start_row
        self._sched_end_rows[scenario_key] = end_row

        _section_header(ws, 4, 1, 9, f'{scenario_label} Overview')
        if schedule:
            total_effort_formula = f'=SUM(H{start_row}:H{end_row})'
            strategic_formula = (
                f'=IFERROR(COUNTIF(G{start_row}:G{end_row},"Strategic")'
                f'/COUNTA(A{start_row}:A{end_row}),0)'
            )
        else:
            total_effort_formula = '=0'
            strategic_formula = '=0'

        _metric_card(
            ws, 5, 1, 3,
            len(schedule),
            'Scheduled Packages',
            '#,##0',
            accent_color,
            'Number of packages in this scenario wave plan',
        )
        _metric_card(
            ws, 5, 4, 6,
            total_effort_formula,
            'Total Effort (days)',
            '#,##0',
            accent_color,
            'Aggregate estimated effort across all scheduled packages',
        )
        _metric_card(
            ws, 5, 7, 9,
            strategic_formula,
            'Strategic Coverage',
            '0.0%',
            accent_color,
            'Share of packages using Strategic mode',
        )

        # Excel Table
        if schedule:
            tbl_name = f"Sched{scenario_key}"
            tbl_ref = f"A{hdr_row}:I{end_row}"
            tbl = Table(displayName=tbl_name, ref=tbl_ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium13", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True,
                showColumnStripes=False)
            ws.add_table(tbl)

        # Conditional formatting — mode colours
        if schedule:
            mode_range = f"G{start_row}:G{end_row}"
            ws.conditional_formatting.add(mode_range, CellIsRule(
                operator='equal', formula=['"Build-to-Legacy"'],
                fill=PatternFill('solid', fgColor=COLORS['build_to_legacy'])))
            ws.conditional_formatting.add(mode_range, CellIsRule(
                operator='equal', formula=['"Bridge-to-Model"'],
                fill=PatternFill('solid', fgColor=COLORS['bridge_to_model'])))
            ws.conditional_formatting.add(mode_range, CellIsRule(
                operator='equal', formula=['"Strategic"'],
                fill=PatternFill('solid', fgColor=COLORS['strategic'])))

        ws.auto_filter.ref = f"A{hdr_row}:I{end_row}"
        ws.freeze_panes = f'A{start_row}'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 6: EXECUTIVE SUMMARY
    # ══════════════════════════════════════════════════════════════════

    def _create_executive_summary(self):
        ws = self.wb.create_sheet('Executive_Summary')
        ws.sheet_properties.tabColor = COLORS['primary']

        for col in range(1, 10):
            ws.column_dimensions[_col_letter(col)].width = 18
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 28

        _sheet_banner(
            ws, 1, 9,
            'Migration Roadmap Executive Summary',
            f'Generated {datetime.now().strftime("%B %d, %Y")} | Steering view of scope, schedule, capacity and scenario trade-offs.',
            COLORS['primary_light'],
        )

        # ── KPI Row ────────────────────────────────────────────────
        row = 4
        _section_header(ws, row, 2, 9, 'Portfolio Snapshot')
        row += 1

        # Total Packages = COUNTA of Package_Data column A minus header
        pkg_sr = self._pkg_data_start_row
        pkg_er = self._pkg_data_end_row

        _metric_card(
            ws, row, 2, 3,
            f"=COUNTA(Package_Data!A{pkg_sr}:A{pkg_er})",
            'Total Packages',
            '#,##0',
            COLORS['primary_light'],
            'Scope of migration packages in the workbook',
        )
        _metric_card(
            ws, row, 4, 5,
            f'={len(self.results)}',
            'Scenarios Modelled',
            '#,##0',
            COLORS['accent_amber'],
            'Fast Exit, Balanced and Target-First',
        )
        _metric_card(
            ws, row, 6, 7,
            f'={self._assumption_cells.get("horizon_months", "60")}',
            'Planning Horizon (months)',
            '#,##0',
            COLORS['accent_green'],
            'Editable in the Assumptions tab',
        )
        _metric_card(
            ws, row, 8, 9,
            f'={self._assumption_cells.get("team_capacity", "6")}',
            'Parallel Capacity',
            '#,##0',
            COLORS['accent_red'],
            'Maximum packages delivered concurrently',
        )

        row += 5

        # ── Scenario Comparison Table ──────────────────────────────
        _section_header(ws, row, 2, 6, 'Scenario Comparison')
        row += 1

        comp_headers = ['Metric'] + SCENARIO_NAMES + ['Best']
        for ci, h in enumerate(comp_headers, 2):
            _hdr_cell(ws, row, ci, h)
        row += 1

        # Metrics rows — using formulas referencing schedule sheets
        metric_rows = []

        # Duration = MAX(End_Month) per schedule
        metric_rows.append(('Duration (months)', [
            f'=MAX(Schedule_{k}!E{self._sched_start_rows.get(k,2)}:'
            f'E{self._sched_end_rows.get(k,2)})'
            for k in SCENARIO_KEYS
        ], '#,##0'))

        # Strategic % = COUNTIF(Mode="Strategic") / COUNTA(IDs)
        metric_rows.append(('Strategic Coverage', [
            f'=IFERROR(COUNTIF(Schedule_{k}!G:G,"Strategic")'
            f'/COUNTA(Schedule_{k}!A{self._sched_start_rows.get(k,2)}:'
            f'A{self._sched_end_rows.get(k,2)}),0)'
            for k in SCENARIO_KEYS
        ], '0.0%'))

        # Build-to-Legacy count
        metric_rows.append(('Build-to-Legacy', [
            f'=COUNTIF(Schedule_{k}!G:G,"Build-to-Legacy")'
            for k in SCENARIO_KEYS
        ], '#,##0'))

        # Bridge-to-Model count
        metric_rows.append(('Bridge-to-Model', [
            f'=COUNTIF(Schedule_{k}!G:G,"Bridge-to-Model")'
            for k in SCENARIO_KEYS
        ], '#,##0'))

        # Strategic count
        metric_rows.append(('Strategic Mode', [
            f'=COUNTIF(Schedule_{k}!G:G,"Strategic")'
            for k in SCENARIO_KEYS
        ], '#,##0'))

        # Total Effort
        metric_rows.append(('Total Effort (days)', [
            f'=SUM(Schedule_{k}!H{self._sched_start_rows.get(k,2)}:'
            f'H{self._sched_end_rows.get(k,2)})'
            for k in SCENARIO_KEYS
        ], '#,##0'))

        dur_row = row
        strategic_row = row + 1

        for mname, formulas, fmt in metric_rows:
            _label_cell(ws, row, 2, mname, bold=True)
            for ci, f in enumerate(formulas, 3):
                _formula_cell(ws, row, ci, f, fmt)

            # "Best" column — MIN for duration/effort, MAX for strategic%
            col_c = _col_letter(3)
            col_e = _col_letter(3 + len(SCENARIO_KEYS) - 1)
            if 'Duration' in mname or 'Effort' in mname or 'Legacy' in mname:
                best_f = f'=MIN({col_c}{row}:{col_e}{row})'
            elif 'Strategic Coverage' in mname or 'Strategic Mode' in mname:
                best_f = f'=MAX({col_c}{row}:{col_e}{row})'
            else:
                best_f = f'=MIN({col_c}{row}:{col_e}{row})'
            _formula_cell(ws, row, 6, best_f, fmt)
            ws.cell(row=row, column=6).fill = FILL_PANEL_ALT
            ws.cell(row=row, column=6).font = Font(
                name='Aptos', size=11, bold=True, color=COLORS['primary']
            )

            row += 1

        # Color scale on duration row
        if len(SCENARIO_KEYS) > 0:
            ws.conditional_formatting.add(
                f'C{dur_row}:E{dur_row}',
                ColorScaleRule(start_type='min', start_color=COLORS['accent_green'],
                               end_type='max', end_color=COLORS['accent_red']))
            ws.conditional_formatting.add(
                f'C{strategic_row}:E{strategic_row}',
                ColorScaleRule(start_type='min', start_color='E7EEF6',
                               end_type='max', end_color='8CC6A3'))

        row += 2

        # ── Recommendations (formula-driven text) ──────────────────
        _section_header(ws, row, 2, 9, 'Recommendations')
        row += 1

        recommendations = self._generate_recommendations()
        for i, rec in enumerate(recommendations[:6], 1):
            _callout_row(
                ws, row, 2, 9,
                f'{i}. {rec}',
                COLORS['accent_green'] if i == 1 else COLORS['primary_light'],
            )
            row += 1

        ws.freeze_panes = 'B5'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 7: SCENARIO COMPARISON
    # ══════════════════════════════════════════════════════════════════

    def _create_scenario_comparison(self):
        ws = self.wb.create_sheet('Scenario_Comparison')
        ws.sheet_properties.tabColor = COLORS['primary_light']

        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 28
        for c in range(3, 8):
            ws.column_dimensions[_col_letter(c)].width = 18
        ws.column_dimensions['H'].width = 16

        _sheet_banner(
            ws, 1, 8,
            'Detailed Scenario Comparison',
            'Expanded side-by-side view of duration, strategic coverage, tech debt and mode mix across the three roadmap options.',
            COLORS['primary_light'],
        )

        row = 4
        _section_header(ws, row, 2, 7, 'Scorecard')
        row += 1

        best_duration_formula = (
            f'=MIN('
            f'MAX(Schedule_FastExit!E{self._sched_start_rows.get("FastExit",2)}:E{self._sched_end_rows.get("FastExit",2)}),'
            f'MAX(Schedule_Balanced!E{self._sched_start_rows.get("Balanced",2)}:E{self._sched_end_rows.get("Balanced",2)}),'
            f'MAX(Schedule_TargetFirst!E{self._sched_start_rows.get("TargetFirst",2)}:E{self._sched_end_rows.get("TargetFirst",2)}))'
        )
        best_strategic_formula = (
            '=MAX('
            f'IFERROR(COUNTIF(Schedule_FastExit!G:G,"Strategic")/COUNTA(Schedule_FastExit!A{self._sched_start_rows.get("FastExit",2)}:A{self._sched_end_rows.get("FastExit",2)}),0),'
            f'IFERROR(COUNTIF(Schedule_Balanced!G:G,"Strategic")/COUNTA(Schedule_Balanced!A{self._sched_start_rows.get("Balanced",2)}:A{self._sched_end_rows.get("Balanced",2)}),0),'
            f'IFERROR(COUNTIF(Schedule_TargetFirst!G:G,"Strategic")/COUNTA(Schedule_TargetFirst!A{self._sched_start_rows.get("TargetFirst",2)}:A{self._sched_end_rows.get("TargetFirst",2)}),0))'
        )
        best_effort_formula = (
            f'=MIN('
            f'SUM(Schedule_FastExit!H{self._sched_start_rows.get("FastExit",2)}:H{self._sched_end_rows.get("FastExit",2)}),'
            f'SUM(Schedule_Balanced!H{self._sched_start_rows.get("Balanced",2)}:H{self._sched_end_rows.get("Balanced",2)}),'
            f'SUM(Schedule_TargetFirst!H{self._sched_start_rows.get("TargetFirst",2)}:H{self._sched_end_rows.get("TargetFirst",2)}))'
        )

        _metric_card(
            ws, row, 2, 3, best_duration_formula, 'Shortest Duration', '#,##0',
            COLORS['accent_green'], 'Lower is better'
        )
        _metric_card(
            ws, row, 4, 5, best_strategic_formula, 'Highest Strategic Coverage', '0.0%',
            COLORS['primary_light'], 'Higher is better'
        )
        _metric_card(
            ws, row, 6, 7, best_effort_formula, 'Lowest Total Effort', '#,##0',
            COLORS['accent_amber'], 'Lower is better'
        )

        row += 5
        _section_header(ws, row, 2, 6, 'Metrics Matrix')
        row += 1

        comp_headers = ['Metric'] + SCENARIO_NAMES + ['Best ✓']
        for ci, h in enumerate(comp_headers, 2):
            _hdr_cell(ws, row, ci, h)
        row += 1

        # Build detailed metrics
        detail_metrics = []

        # Duration
        detail_metrics.append(('Total Duration (months)', [
            f'=MAX(Schedule_{k}!E{self._sched_start_rows.get(k,2)}:'
            f'E{self._sched_end_rows.get(k,2)})'
            for k in SCENARIO_KEYS
        ], '#,##0', 'min'))

        # Strategic %
        detail_metrics.append(('Strategic Coverage (%)', [
            f'=IFERROR(COUNTIF(Schedule_{k}!G:G,"Strategic")'
            f'/COUNTA(Schedule_{k}!A{self._sched_start_rows.get(k,2)}:'
            f'A{self._sched_end_rows.get(k,2)}),0)'
            for k in SCENARIO_KEYS
        ], '0.0%', 'max'))

        # Tech Debt Score (SUMPRODUCT with mode penalty)
        for k in SCENARIO_KEYS:
            sr = self._sched_start_rows.get(k, 2)
            er = self._sched_end_rows.get(k, 2)

        detail_metrics.append(('Tech Debt Score', [
            f'=COUNTIF(Schedule_{k}!G:G,"Build-to-Legacy")*3'
            f'+COUNTIF(Schedule_{k}!G:G,"Bridge-to-Model")*1.5'
            f'+COUNTIF(Schedule_{k}!G:G,"Strategic")*0.5'
            for k in SCENARIO_KEYS
        ], '#,##0.0', 'min'))

        # Mode counts
        for mode_label in MODE_NAMES:
            detail_metrics.append((f'{mode_label} Count', [
                f'=COUNTIF(Schedule_{k}!G:G,"{mode_label}")'
                for k in SCENARIO_KEYS
            ], '#,##0', 'min' if 'Legacy' in mode_label else 'max'))

        # Total effort
        detail_metrics.append(('Total Effort (days)', [
            f'=SUM(Schedule_{k}!H{self._sched_start_rows.get(k,2)}:'
            f'H{self._sched_end_rows.get(k,2)})'
            for k in SCENARIO_KEYS
        ], '#,##0', 'min'))

        data_start_row = row
        duration_row = row
        strategic_row = row + 1
        for mname, formulas, fmt, best_dir in detail_metrics:
            _label_cell(ws, row, 2, mname, bold=True)
            for ci, f in enumerate(formulas, 3):
                _formula_cell(ws, row, ci, f, fmt)

            # Best indicator
            col_c = _col_letter(3)
            col_e = _col_letter(3 + len(SCENARIO_KEYS) - 1)
            if best_dir == 'min':
                best_f = (f'=IF(MIN({col_c}{row}:{col_e}{row})>0,'
                          f'IF(MIN({col_c}{row}:{col_e}{row})='
                          f'MIN({col_c}{row}:{col_e}{row}),"✓",""),"")')
                # Simpler: show the min value
                best_f = f'=MIN({col_c}{row}:{col_e}{row})'
            else:
                best_f = f'=MAX({col_c}{row}:{col_e}{row})'
            _formula_cell(ws, row, 6, best_f, fmt)
            ws.cell(row=row, column=6).fill = FILL_PANEL_ALT
            ws.cell(row=row, column=6).font = Font(
                name='Aptos', size=11, bold=True, color=COLORS['primary']
            )
            row += 1

        data_end_row = row - 1

        ws.conditional_formatting.add(
            f'C{duration_row}:E{duration_row}',
            ColorScaleRule(start_type='min', start_color=COLORS['accent_green'],
                           end_type='max', end_color=COLORS['accent_red'])
        )
        ws.conditional_formatting.add(
            f'C{strategic_row}:E{strategic_row}',
            ColorScaleRule(start_type='min', start_color='E7EEF6',
                           end_type='max', end_color='8CC6A3')
        )

        row += 2

        # ── Stacked bar chart for mode distribution ────────────────
        _section_header(ws, row, 2, 6, 'Mode Distribution')
        row += 1
        chart_data_start = row

        _hdr_cell(ws, row, 2, 'Mode')
        for ci, sn in enumerate(SCENARIO_NAMES, 3):
            _hdr_cell(ws, row, ci, sn)
        row += 1

        for mode_label in MODE_NAMES:
            _label_cell(ws, row, 2, mode_label, bold=True)
            for ci, k in enumerate(SCENARIO_KEYS, 3):
                _formula_cell(ws, row, ci,
                              f'=COUNTIF(Schedule_{k}!G:G,"{mode_label}")',
                              '#,##0')
            row += 1
        chart_data_end = row - 1

        # Build chart
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = "Packages by Migration Mode"
        chart.style = 11
        chart.y_axis.title = "Package Count"

        data = Reference(ws, min_col=3, min_row=chart_data_start,
                         max_col=2 + len(SCENARIO_KEYS),
                         max_row=chart_data_end)
        cats = Reference(ws, min_col=2, min_row=chart_data_start + 1,
                         max_row=chart_data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 12

        ws.add_chart(chart, f'B{row + 1}')

        ws.freeze_panes = 'C10'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 8: TIMELINE GANTT
    # ══════════════════════════════════════════════════════════════════

    def _create_gantt_timeline(self):
        ws = self.wb.create_sheet('Timeline_Gantt')
        ws.sheet_properties.tabColor = COLORS['secondary']

        # Use balanced schedule (middle scenario) for the Gantt
        gantt_key = 'Balanced' if 'Balanced' in self.scenario_schedules else SCENARIO_KEYS[0]
        schedule = self.scenario_schedules.get(gantt_key, [])

        if not schedule:
            _sheet_banner(
                ws, 1, 4,
                'Timeline Gantt',
                'No schedule data available for the selected scenario.',
                COLORS['primary_light'],
            )
            ws.cell(row=4, column=1, value='No schedule data available').font = FONT_FORMULA
            return

        max_month = max(item.get('end_month', item.get('end', 0)) for item in schedule)
        max_month = min(max_month, 60)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 9
        for mi in range(max_month):
            ws.column_dimensions[_col_letter(3 + mi)].width = 3.5

        _sheet_banner(
            ws, 1, 2 + max_month,
            'Timeline Gantt',
            'Balanced scenario view grouped by quarter to make wave transitions and overlaps easier to read.',
            COLORS['primary_light'],
        )

        ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
        ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
        pkg_hdr = ws.cell(row=4, column=1, value='Package')
        mode_hdr = ws.cell(row=4, column=2, value='Mode')
        for cell in (pkg_hdr, mode_hdr):
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        ws.cell(row=5, column=1).fill = FILL_HEADER
        ws.cell(row=5, column=2).fill = FILL_HEADER
        ws.cell(row=5, column=1).border = THIN_BORDER
        ws.cell(row=5, column=2).border = THIN_BORDER

        for mi in range(max_month):
            month_col = 3 + mi
            quarter_number = (mi // 3) + 1
            if mi % 3 == 0:
                quarter_end = min(month_col + 2, 2 + max_month)
                ws.merge_cells(start_row=4, start_column=month_col,
                               end_row=4, end_column=quarter_end)
                q_cell = ws.cell(row=4, column=month_col, value=f'Q{quarter_number}')
                q_cell.font = Font(name='Aptos', size=9, bold=True, color=COLORS['primary'])
                q_cell.fill = PatternFill('solid', fgColor=COLORS['quarter_fill'])
                q_cell.alignment = ALIGN_CENTER
                q_cell.border = THIN_BORDER
                for ci in range(month_col + 1, quarter_end + 1):
                    ws.cell(row=4, column=ci).fill = PatternFill(
                        'solid', fgColor=COLORS['quarter_fill']
                    )
                    ws.cell(row=4, column=ci).border = THIN_BORDER

            c = ws.cell(row=5, column=month_col, value=f'M{mi + 1:02d}')
            c.font = Font(name='Aptos', size=8, bold=True, color=COLORS['white'])
            c.fill = FILL_HEADER
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER

        data_start_row = 6
        for ri, item in enumerate(schedule, data_start_row):
            name = item.get('package_name', item.get('name', ''))
            mode = item.get('selected_mode', item.get('mode', ''))
            mode_display = mode.replace('_', '-').title()
            if mode_display == 'Build-To-Legacy':
                mode_display = 'Build-to-Legacy'
            elif mode_display == 'Bridge-To-Model':
                mode_display = 'Bridge-to-Model'

            base_fill = FILL_PANEL if (ri - data_start_row) % 2 == 0 else FILL_PANEL_ALT

            name_cell = ws.cell(row=ri, column=1, value=name[:30])
            name_cell.font = FONT_FORMULA
            name_cell.border = THIN_BORDER
            name_cell.fill = base_fill
            name_cell.alignment = ALIGN_LEFT

            abbrev = {'Build-to-Legacy': 'BtL', 'Bridge-to-Model': 'BtM',
                      'Strategic': 'Str'}.get(mode_display, mode_display[:3])
            mode_cell = ws.cell(row=ri, column=2, value=abbrev)
            mode_cell.font = Font(name='Aptos', size=8, bold=True, color=COLORS['secondary'])
            mode_cell.alignment = ALIGN_CENTER
            mode_cell.border = THIN_BORDER
            mode_cell.fill = base_fill

            start = item.get('start_month', 0)
            end = item.get('end_month', 0)

            # Map mode → fill colour
            fill_map = {
                'Build-to-Legacy': PatternFill('solid', fgColor=COLORS['mode_btl']),
                'Bridge-to-Model': PatternFill('solid', fgColor=COLORS['mode_btm']),
                'Strategic':       PatternFill('solid', fgColor=COLORS['mode_str']),
            }
            fill = fill_map.get(mode_display,
                                PatternFill('solid', fgColor='CCCCCC'))

            for mi in range(max_month):
                cell = ws.cell(row=ri, column=3 + mi)
                cell.border = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0'))
                cell.fill = base_fill
                if start <= mi < end:
                    cell.fill = fill

        # Legend
        legend_row = data_start_row + len(schedule) + 2
        _section_header(ws, legend_row, 1, 4, 'Mode Legend')
        for i, (mode_l, color_key) in enumerate([
            ('Build-to-Legacy', 'mode_btl'),
            ('Bridge-to-Model', 'mode_btm'),
            ('Strategic',       'mode_str'),
        ]):
            c = ws.cell(row=legend_row + 1 + i, column=1, value=mode_l)
            c.fill = PatternFill('solid', fgColor=COLORS[color_key])
            c.font = Font(name='Aptos', size=10, bold=True, color=COLORS['white'])
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER
            desc = ws.cell(row=legend_row + 1 + i, column=2,
                           value='Timeline bar colour')
            desc.font = FONT_LABEL
            desc.fill = FILL_PANEL
            desc.border = THIN_BORDER

        ws.freeze_panes = 'C6'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 9: EFFORT & COST MODEL (NEW)
    # ══════════════════════════════════════════════════════════════════

    def _create_effort_cost_model(self):
        ws = self.wb.create_sheet('Effort_Cost_Model')
        ws.sheet_properties.tabColor = COLORS['accent_green']

        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 30
        for c in range(3, 8):
            ws.column_dimensions[_col_letter(c)].width = 18

        _sheet_banner(
            ws, 1, 7,
            'Effort & Cost Model',
            'Budget and burn-rate views linked directly to the editable Assumptions tab.',
            COLORS['accent_green'],
        )

        row = 4

        # ── Per-scenario cost summary ──────────────────────────────
        _section_header(ws, row, 2, 7, 'Scenario Cost Summary')
        row += 1

        cost_headers = ['Metric'] + SCENARIO_NAMES + ['Delta (max−min)']
        for ci, h in enumerate(cost_headers, 2):
            _hdr_cell(ws, row, ci, h)
        row += 1

        fte_rate_ref = self._assumption_cells.get('fte_daily_rate',
                                                   'Assumptions!$C$7')

        # Total Effort
        effort_formulas = []
        for k in SCENARIO_KEYS:
            sr = self._sched_start_rows.get(k, 2)
            er = self._sched_end_rows.get(k, 2)
            effort_formulas.append(
                f'=SUM(Schedule_{k}!H{sr}:H{er})')

        _label_cell(ws, row, 2, 'Total Effort (days)', bold=True)
        effort_cells = []
        for ci, f in enumerate(effort_formulas, 3):
            _formula_cell(ws, row, ci, f, '#,##0')
            effort_cells.append(f'{_col_letter(ci)}{row}')
        _formula_cell(ws, row, 6,
                      f'=MAX({effort_cells[0]}:{effort_cells[-1]})'
                      f'-MIN({effort_cells[0]}:{effort_cells[-1]})', '#,##0')
        effort_row = row
        row += 1

        # Total Cost = Effort × FTE rate
        _label_cell(ws, row, 2, 'Total Cost ($)', bold=True)
        cost_cells = []
        for ci in range(3, 3 + len(SCENARIO_KEYS)):
            cost_f = f'={_col_letter(ci)}{effort_row}*{fte_rate_ref}'
            _formula_cell(ws, row, ci, cost_f, '$#,##0')
            cost_cells.append(f'{_col_letter(ci)}{row}')
        _formula_cell(ws, row, 6,
                      f'=MAX({cost_cells[0]}:{cost_cells[-1]})'
                      f'-MIN({cost_cells[0]}:{cost_cells[-1]})', '$#,##0')
        cost_row = row
        row += 1

        # Duration
        _label_cell(ws, row, 2, 'Duration (months)', bold=True)
        dur_cells = []
        for ci, k in enumerate(SCENARIO_KEYS, 3):
            sr = self._sched_start_rows.get(k, 2)
            er = self._sched_end_rows.get(k, 2)
            _formula_cell(ws, row, ci,
                          f'=MAX(Schedule_{k}!E{sr}:E{er})', '#,##0')
            dur_cells.append(f'{_col_letter(ci)}{row}')
        _formula_cell(ws, row, 6,
                      f'=MAX({dur_cells[0]}:{dur_cells[-1]})'
                      f'-MIN({dur_cells[0]}:{dur_cells[-1]})', '#,##0')
        row += 1

        # Cost per Month
        _label_cell(ws, row, 2, 'Avg Cost / Month ($)', bold=True)
        for ci in range(3, 3 + len(SCENARIO_KEYS)):
            _formula_cell(ws, row, ci,
                          f'=IFERROR({_col_letter(ci)}{cost_row}'
                          f'/{_col_letter(ci)}{cost_row+1},0)', '$#,##0')
        ws.cell(row=effort_row, column=6).fill = FILL_PANEL_ALT
        ws.cell(row=cost_row, column=6).fill = FILL_PANEL_ALT
        ws.cell(row=cost_row + 1, column=6).fill = FILL_PANEL_ALT
        row += 2

        # ── Monthly Burn Rate ──────────────────────────────────────
        _section_header(ws, row, 2, 7, 'Monthly Burn Rate (active effort × daily rate)')
        row += 1
        burn_hdr_row = row

        _hdr_cell(ws, row, 2, 'Month')
        for ci, sn in enumerate(SCENARIO_NAMES, 3):
            _hdr_cell(ws, row, ci, sn)
        row += 1

        burn_start = row
        max_months_display = 36  # Show up to 36 months

        for m in range(max_months_display):
            _label_cell(ws, row, 2, f'M{m}')
            for ci, k in enumerate(SCENARIO_KEYS, 3):
                sr = self._sched_start_rows.get(k, 2)
                er = self._sched_end_rows.get(k, 2)
                # SUMPRODUCT: sum effort of packages active in month m,
                # divided by their duration, times daily rate
                # Simplified: count active packages × average monthly effort × rate
                formula = (
                    f'=SUMPRODUCT('
                    f'(Schedule_{k}!D{sr}:D{er}<={m})*'
                    f'(Schedule_{k}!E{sr}:E{er}>{m})*'
                    f'Schedule_{k}!H{sr}:H{er}'
                    f'/IF(Schedule_{k}!F{sr}:F{er}=0,1,'
                    f'Schedule_{k}!F{sr}:F{er}))'
                    f'*{fte_rate_ref}/20'
                )
                _formula_cell(ws, row, ci, formula, '$#,##0')
            row += 1
        burn_end = row - 1

        # Line chart for burn rate
        chart = LineChart()
        chart.title = "Monthly Burn Rate by Scenario"
        chart.style = 11
        chart.y_axis.title = "Cost ($)"
        chart.x_axis.title = "Month"
        chart.width = 20
        chart.height = 12

        data = Reference(ws, min_col=3, min_row=burn_hdr_row,
                         max_col=2 + len(SCENARIO_KEYS),
                         max_row=burn_end)
        cats = Reference(ws, min_col=2, min_row=burn_start,
                         max_row=burn_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, f'B{row + 1}')

        ws.freeze_panes = 'C5'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 10: RISK ASSESSMENT (NEW)
    # ══════════════════════════════════════════════════════════════════

    def _create_risk_assessment(self):
        ws = self.wb.create_sheet('Risk_Assessment')
        ws.sheet_properties.tabColor = COLORS['accent_red']

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 14

        _sheet_banner(
            ws, 1, 9,
            'Risk Assessment Model',
            'Weighted risk register with conditional formatting, mitigation prioritisation and duration uplift estimates.',
            COLORS['accent_red'],
        )
        _section_header(ws, 4, 1, 9, 'Risk Register')

        # Headers
        headers = ['Package_ID', 'Name', 'Domain', 'Complexity',
                   'Dependencies', 'Business_Value', 'Weighted_Risk',
                   'Risk_Category', 'Risk_Adj_Duration']

        hdr_row = 5
        for ci, h in enumerate(headers, 1):
            _hdr_cell(ws, hdr_row, ci, h)

        # Risk weight references
        w_comp = self._assumption_cells.get('risk_w_complexity',
                                             'Assumptions!$C$40')
        w_dep  = self._assumption_cells.get('risk_w_dependency',
                                             'Assumptions!$C$41')
        w_bv   = self._assumption_cells.get('risk_w_bv',
                                             'Assumptions!$C$42')
        w_vol  = self._assumption_cells.get('risk_w_volume',
                                             'Assumptions!$C$43')
        risk_buffer = self._assumption_cells.get('risk_buffer_factor',
                                                  'Assumptions!$C$9')

        pkg_sr = self._pkg_data_start_row
        start_row = hdr_row + 1

        for ri, pkg in enumerate(self.packages, start_row):
            pkg_row = pkg_sr + ri - start_row  # corresponding row in Package_Data

            # Package ID and Name from Package_Data via formula
            _formula_cell(ws, ri, 1,
                          f'=Package_Data!A{pkg_row}')
            _formula_cell(ws, ri, 2,
                          f'=Package_Data!B{pkg_row}')
            ws.cell(row=ri, column=2).alignment = ALIGN_LEFT
            _formula_cell(ws, ri, 3,
                          f'=Package_Data!C{pkg_row}')

            # Complexity from Package_Data
            _formula_cell(ws, ri, 4,
                          f'=Package_Data!F{pkg_row}', '0.0')

            # Dependencies (upstream + downstream)
            _formula_cell(ws, ri, 5,
                          f'=Package_Data!I{pkg_row}+Package_Data!J{pkg_row}',
                          '#,##0')

            # Business Value
            _formula_cell(ws, ri, 6,
                          f'=Package_Data!G{pkg_row}', '#,##0')

            # Weighted Risk Score (normalised 0-1 scale)
            # = (Complexity/5 * W1 + Deps/10 * W2 + (1-BV/10) * W3 + Jobs/50 * W4)
            risk_formula = (
                f'=MIN(1,'
                f'(D{ri}/5)*{w_comp}'
                f'+(E{ri}/10)*{w_dep}'
                f'+((10-F{ri})/10)*{w_bv}'
                f'+(Package_Data!D{pkg_row}/50)*{w_vol})'
            )
            _formula_cell(ws, ri, 7, risk_formula, '0.00')

            # Risk Category
            cat_formula = (
                f'=IF(G{ri}>0.7,"Critical",'
                f'IF(G{ri}>0.4,"High",'
                f'IF(G{ri}>0.2,"Medium","Low")))'
            )
            _formula_cell(ws, ri, 8, cat_formula)

            # Risk-Adjusted Duration  (using Balanced schedule as base)
            bal_key = 'Balanced'
            bal_sr = self._sched_start_rows.get(bal_key, 2)
            bal_er = self._sched_end_rows.get(bal_key, 2)
            # VLOOKUP the duration for this package from balanced schedule
            risk_dur_formula = (
                f'=IFERROR(VLOOKUP(A{ri},'
                f'Schedule_{bal_key}!A{bal_sr}:F{bal_er},'
                f'6,FALSE)*(1+G{ri}*{risk_buffer}),"-")'
            )
            _formula_cell(ws, ri, 9, risk_dur_formula, '0.0')

            for ci in range(1, 10):
                ws.cell(row=ri, column=ci).border = THIN_BORDER

        end_row = start_row + self.n_packages - 1

        # Conditional formatting — risk heat map on weighted risk
        if self.n_packages > 0:
            ws.conditional_formatting.add(
                f'G{start_row}:G{end_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='28A745',
                    mid_type='num', mid_value=0.4, mid_color='FFC107',
                    end_type='num', end_value=1.0, end_color='DC3545'))

            # Risk category conditional formatting
            cat_range = f'H{start_row}:H{end_row}'
            ws.conditional_formatting.add(cat_range, CellIsRule(
                operator='equal', formula=['"Critical"'],
                fill=PatternFill('solid', fgColor='F8D7DA'),
                font=Font(color='721C24', bold=True)))
            ws.conditional_formatting.add(cat_range, CellIsRule(
                operator='equal', formula=['"High"'],
                fill=PatternFill('solid', fgColor='FFF3CD'),
                font=Font(color='856404', bold=True)))
            ws.conditional_formatting.add(cat_range, CellIsRule(
                operator='equal', formula=['"Medium"'],
                fill=PatternFill('solid', fgColor='D4EDDA'),
                font=Font(color='155724')))
            ws.conditional_formatting.add(cat_range, CellIsRule(
                operator='equal', formula=['"Low"'],
                fill=PatternFill('solid', fgColor='D1ECF1'),
                font=Font(color='0C5460')))

        # ── Summary Section ────────────────────────────────────────
        summary_row = end_row + 3
        _section_header(ws, summary_row, 1, 9, 'Risk Summary')
        summary_row += 1

        risk_categories = ['Critical', 'High', 'Medium', 'Low']
        for cat in risk_categories:
            _label_cell(ws, summary_row, 1, f'{cat} Packages', bold=True)
            _formula_cell(ws, summary_row, 2,
                          f'=COUNTIF(H{start_row}:H{end_row},"{cat}")',
                          '#,##0')
            _formula_cell(ws, summary_row, 3,
                          f'=IFERROR(COUNTIF(H{start_row}:H{end_row},"{cat}")'
                          f'/COUNTA(H{start_row}:H{end_row}),0)', '0.0%')
            summary_row += 1

        # Average risk score
        _label_cell(ws, summary_row, 1, 'Average Risk Score', bold=True)
        _formula_cell(ws, summary_row, 2,
                      f'=AVERAGE(G{start_row}:G{end_row})', '0.00')

        # Pie chart for risk distribution
        pie_chart = PieChart()
        pie_chart.title = "Risk Distribution"
        pie_chart.style = 10
        pie_chart.width = 14
        pie_chart.height = 10
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True
        pie_chart.dataLabels.showLeaderLines = True

        data_ref = Reference(ws, min_col=2, min_row=end_row + 4,
                             max_row=end_row + 7)
        cat_ref = Reference(ws, min_col=1, min_row=end_row + 4,
                            max_row=end_row + 7)
        pie_chart.add_data(data_ref)
        pie_chart.set_categories(cat_ref)

        ws.add_chart(pie_chart, f'E{end_row + 3}')

        ws.auto_filter.ref = f"A{hdr_row}:I{end_row}"
        ws.freeze_panes = 'A6'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 11: DEPENDENCY ANALYSIS (NEW)
    # ══════════════════════════════════════════════════════════════════

    def _create_dependency_analysis(self):
        ws = self.wb.create_sheet('Dependency_Analysis')
        ws.sheet_properties.tabColor = COLORS['mode_btm']

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 16

        _sheet_banner(
            ws, 1, 8,
            'Dependency & Impact Analysis',
            'Critical-path proxy, dependency load and domain-level impact summary for migration sequencing.',
            COLORS['mode_btm'],
        )
        _section_header(ws, 4, 1, 8, 'Dependency Register')

        # Headers
        headers = ['Package_ID', 'Name', 'Upstream', 'Downstream',
                   'Total_Deps', 'Fan_Out_Score', 'Critical_Path',
                   'Impact_Score']
        hdr_row = 5
        for ci, h in enumerate(headers, 1):
            _hdr_cell(ws, hdr_row, ci, h)

        pkg_sr = self._pkg_data_start_row
        dep_threshold_ref = self._assumption_cells.get(
            'dep_crit_threshold', 'Assumptions!$C$10')
        complexity_threshold_ref = self._assumption_cells.get(
            'complexity_crit_threshold', 'Assumptions!$C$11')

        start_row = hdr_row + 1
        for ri, pkg in enumerate(self.packages, start_row):
            pkg_row = pkg_sr + ri - start_row

            # Package ID, Name from Package_Data
            _formula_cell(ws, ri, 1, f'=Package_Data!A{pkg_row}')
            _formula_cell(ws, ri, 2, f'=Package_Data!B{pkg_row}')
            ws.cell(row=ri, column=2).alignment = ALIGN_LEFT

            # Upstream Count
            _formula_cell(ws, ri, 3,
                          f'=Package_Data!I{pkg_row}', '#,##0')

            # Downstream Count
            _formula_cell(ws, ri, 4,
                          f'=Package_Data!J{pkg_row}', '#,##0')

            # Total Dependencies
            _formula_cell(ws, ri, 5, f'=C{ri}+D{ri}', '#,##0')

            # Fan-Out Score = Downstream × (1 + Complexity/5)
            _formula_cell(ws, ri, 6,
                          f'=D{ri}*(1+Package_Data!F{pkg_row}/5)', '0.0')

            # Critical Path indicator
            crit_formula = (
                f'=IF(AND(D{ri}>={dep_threshold_ref},'
                f'Package_Data!F{pkg_row}>={complexity_threshold_ref}),'
                f'"Critical Path","")'
            )
            _formula_cell(ws, ri, 7, crit_formula)

            # Impact Score = Downstream × Business Value / 10
            _formula_cell(ws, ri, 8,
                          f'=D{ri}*Package_Data!G{pkg_row}/10', '0.0')

            for ci in range(1, 9):
                ws.cell(row=ri, column=ci).border = THIN_BORDER

        end_row = start_row + self.n_packages - 1

        # Conditional formatting on Fan-Out Score
        if self.n_packages > 0:
            ws.conditional_formatting.add(
                f'F{start_row}:F{end_row}',
                DataBarRule(start_type='min', end_type='max',
                            color='7986CB', showValue=True,
                            minLength=None, maxLength=None))

            # Highlight critical path items
            ws.conditional_formatting.add(
                f'G{start_row}:G{end_row}',
                CellIsRule(operator='equal',
                           formula=['"Critical Path"'],
                           fill=PatternFill('solid', fgColor='FFCDD2'),
                           font=Font(color='C62828', bold=True)))

            # Color scale on Impact Score
            ws.conditional_formatting.add(
                f'H{start_row}:H{end_row}',
                ColorScaleRule(start_type='min', start_color='FFFFFF',
                               end_type='max', end_color='7986CB'))

        # ── Domain Interconnection Matrix ──────────────────────────
        domains = sorted({p.get('domain', 'Unknown') for p in self.packages})
        if domains:
            matrix_row = end_row + 3
            _section_header(ws, matrix_row, 1, 1 + len(domains),
                            'Domain Dependency Summary')
            matrix_row += 1

            _hdr_cell(ws, matrix_row, 1, 'Domain')
            for ci, d in enumerate(domains, 2):
                _hdr_cell(ws, matrix_row, ci, d)
            matrix_row += 1

            # For each domain pair, count packages with dependencies
            # Since we don't have explicit cross-domain deps, we use
            # a count of packages per domain as a proxy
            for di, domain in enumerate(domains):
                _label_cell(ws, matrix_row + di, 1, domain, bold=True)
                for dj, domain2 in enumerate(domains, 2):
                    if domain == domains[dj - 2]:
                        # Same domain — count internal dependencies
                        count = sum(
                            1 for p in self.packages
                            if p.get('domain', '') == domain
                            and (p.get('upstream_count',
                                       len(p.get('upstream_packages', []))) > 0))
                        _formula_cell(ws, matrix_row + di, dj, count, '#,##0')
                    else:
                        # Cross-domain — estimate
                        _formula_cell(ws, matrix_row + di, dj, '-')
                    ws.cell(row=matrix_row + di, column=dj).border = THIN_BORDER

        # ── Summary stats ──────────────────────────────────────────
        sum_row = (matrix_row + len(domains) + 2) if domains else (end_row + 3)
        _section_header(ws, sum_row, 1, 4, 'Summary Statistics')
        sum_row += 1

        _label_cell(ws, sum_row, 1, 'Total Critical Path Items', bold=True)
        _formula_cell(ws, sum_row, 2,
                      f'=COUNTIF(G{start_row}:G{end_row},"Critical Path")',
                      '#,##0')
        sum_row += 1

        _label_cell(ws, sum_row, 1, 'Max Fan-Out Score', bold=True)
        _formula_cell(ws, sum_row, 2,
                      f'=MAX(F{start_row}:F{end_row})', '0.0')
        sum_row += 1

        _label_cell(ws, sum_row, 1, 'Avg Dependencies per Package', bold=True)
        _formula_cell(ws, sum_row, 2,
                      f'=AVERAGE(E{start_row}:E{end_row})', '0.0')
        sum_row += 1

        _label_cell(ws, sum_row, 1, 'Highest Impact Package', bold=True)
        _formula_cell(ws, sum_row, 2,
                      f'=INDEX(B{start_row}:B{end_row},'
                      f'MATCH(MAX(H{start_row}:H{end_row}),'
                      f'H{start_row}:H{end_row},0))')

        ws.auto_filter.ref = f"A{hdr_row}:H{end_row}"
        ws.freeze_panes = 'A6'

    # ══════════════════════════════════════════════════════════════════
    # SHEET 12: DATA APPENDIX
    # ══════════════════════════════════════════════════════════════════

    def _create_data_appendix(self):
        ws = self.wb.create_sheet('Data_Appendix')
        ws.sheet_properties.tabColor = COLORS['text_muted']

        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18

        _sheet_banner(
            ws, 1, 4,
            'Data Appendix & Configuration',
            'Reference tab for workbook settings, domain totals, mode mix and generation metadata.',
            COLORS['text_muted'],
        )

        row = 4

        # ── Configuration Recap ────────────────────────────────────
        _section_header(ws, row, 2, 4, 'Configuration (from Assumptions)')
        row += 1

        config_items = [
            ('Migration Horizon',
             f'={self._assumption_cells.get("horizon_months", 60)}',
             'months'),
            ('Team Capacity',
             f'={self._assumption_cells.get("team_capacity", 6)}',
             'parallel packages'),
            ('FTE Daily Rate',
             f'={self._assumption_cells.get("fte_daily_rate", 850)}',
             '$'),
            ('Total Packages',
             f'=COUNTA(Package_Data!A{self._pkg_data_start_row}:'
             f'A{self._pkg_data_end_row})',
             'packages'),
            ('Scenarios',
             f'={len(self.results)}',
             'scenarios'),
        ]

        for label, formula, unit in config_items:
            _label_cell(ws, row, 2, label, bold=True)
            _formula_cell(ws, row, 3, formula, '#,##0')
            _label_cell(ws, row, 4, unit)
            row += 1

        row += 1

        # ── Domain Breakdown (formula-driven) ──────────────────────
        _section_header(ws, row, 2, 4, 'Packages by Domain')
        row += 1

        _hdr_cell(ws, row, 2, 'Domain')
        _hdr_cell(ws, row, 3, 'Package Count')
        _hdr_cell(ws, row, 4, 'Total Effort (days)')
        row += 1

        domains = sorted({p.get('domain', 'Unknown') for p in self.packages})
        for domain in domains:
            _label_cell(ws, row, 2, domain, bold=True)
            _formula_cell(ws, row, 3,
                          f'=COUNTIF(Package_Data!C{self._pkg_data_start_row}:'
                          f'C{self._pkg_data_end_row},"{domain}")', '#,##0')
            _formula_cell(ws, row, 4,
                          f'=SUMIF(Package_Data!C{self._pkg_data_start_row}:'
                          f'C{self._pkg_data_end_row},"{domain}",'
                          f'Package_Data!E{self._pkg_data_start_row}:'
                          f'E{self._pkg_data_end_row})', '#,##0')
            row += 1

        # Total row
        _label_cell(ws, row, 2, 'TOTAL', bold=True)
        _formula_cell(ws, row, 3,
                      f'=COUNTA(Package_Data!A{self._pkg_data_start_row}:'
                      f'A{self._pkg_data_end_row})', '#,##0')
        _formula_cell(ws, row, 4,
                      f'=SUM(Package_Data!E{self._pkg_data_start_row}:'
                      f'E{self._pkg_data_end_row})', '#,##0')
        row += 2

        # ── Mode Distribution per Scenario ─────────────────────────
        _section_header(ws, row, 2, 4, 'Mode Distribution per Scenario')
        row += 1

        for sn, k in zip(SCENARIO_NAMES, SCENARIO_KEYS):
            _label_cell(ws, row, 2, sn, bold=True)
            row += 1
            for mode_label in MODE_NAMES:
                _label_cell(ws, row, 2, f'  {mode_label}')
                _formula_cell(ws, row, 3,
                              f'=COUNTIF(Schedule_{k}!G:G,"{mode_label}")',
                              '#,##0')
                row += 1
            row += 1

        # ── Report Metadata ────────────────────────────────────────
        _section_header(ws, row, 2, 4, 'Report Metadata')
        row += 1

        metadata = [
            ('Generated On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Generator Version', '2.0.0 (Financial Model)'),
            ('Report Format', 'Formula-Driven 12-Sheet Workbook'),
            ('Convention', 'Blue = Input, Black = Formula'),
        ]
        for label, value in metadata:
            _label_cell(ws, row, 2, label, bold=True)
            ws.cell(row=row, column=3, value=str(value)).font = FONT_FORMULA
            ws.cell(row=row, column=3).border = THIN_BORDER
            row += 1

        ws.freeze_panes = 'C4'

    # ══════════════════════════════════════════════════════════════════
    # RECOMMENDATION ENGINE
    # ══════════════════════════════════════════════════════════════════

    def _generate_recommendations(self) -> List[str]:
        """Generate strategic recommendations from optimisation results."""
        recs = []

        if not self.results:
            return ["No optimisation results available for analysis."]

        fastest = min(self.results,
                      key=lambda x: x.get('total_duration_months', 999))
        most_strategic = max(self.results,
                             key=lambda x: x.get('strategic_coverage', 0))
        lowest_debt = min(self.results,
                          key=lambda x: x.get('technical_debt_score', 999))

        recs.append(
            f"'{fastest.get('scenario_name', 'Fast Exit')}' achieves the "
            f"shortest timeline at "
            f"{fastest.get('total_duration_months', 0)} months.")

        if most_strategic.get('strategic_coverage', 0) > 0.3:
            recs.append(
                f"'{most_strategic.get('scenario_name', 'Target-First')}' "
                f"maximises strategic alignment at "
                f"{most_strategic.get('strategic_coverage', 0):.0%} coverage.")

        if fastest != lowest_debt:
            delta = (fastest.get('technical_debt_score', 0)
                     - lowest_debt.get('technical_debt_score', 0))
            recs.append(
                f"'{lowest_debt.get('scenario_name', 'Target-First')}' "
                f"reduces tech debt by {delta:.0f} points vs. the fastest "
                f"option.")

        balanced = next(
            (r for r in self.results
             if 'balanced' in r.get('scenario_name', '').lower()), None)
        if balanced:
            recs.append(
                f"The Balanced scenario offers a middle-ground: "
                f"{balanced.get('total_duration_months', 0)} months duration, "
                f"{balanced.get('strategic_coverage', 0):.0%} strategic.")

        team = self.config.get('team_capacity', 6)
        n = len(self.packages)
        recs.append(
            f"With {team} parallel slots handling {n} packages, "
            f"focus early waves on high-value domains for maximum impact.")

        recs.append(
            "Review the Effort & Cost Model sheet for budget projections, "
            "and Risk Assessment for critical packages requiring mitigation.")

        return recs


# ════════════════════════════════════════════════════════════════════════════
# PUBLIC API
# ════════════════════════════════════════════════════════════════════════════

def generate_migration_report(
    optimization_results: List[Dict],
    packages: List[Dict],
    config: Dict,
    output_path: str,
) -> str:
    """
    Generate a professional, formula-driven Excel workbook.

    Args:
        optimization_results: List of scenario results from the optimiser
        packages: List of migration packages
        config: Optimisation configuration
        output_path: Where to save the .xlsx file

    Returns:
        Path to the generated file
    """
    gen = ExcelReportGenerator(optimization_results, packages, config)
    return gen.generate_report(output_path)


def main():
    """CLI entry point."""
    import sys

    if len(sys.argv) < 4:
        print("Usage: python excel_generator.py <results.json> "
              "<packages.json> <config.json> [output.xlsx]")
        sys.exit(1)

    results_file = sys.argv[1]
    packages_file = sys.argv[2]
    config_file = sys.argv[3]
    output_file = (sys.argv[4] if len(sys.argv) > 4
                   else "migration_roadmap_report.xlsx")

    with open(results_file) as f:
        results = json.load(f)
    with open(packages_file) as f:
        packages = json.load(f)
    with open(config_file) as f:
        config = json.load(f)

    if isinstance(results, dict):
        results = [results]

    generate_migration_report(results, packages, config, output_file)
    print(f"\nReport generated: {output_file}")


if __name__ == "__main__":
    main()
